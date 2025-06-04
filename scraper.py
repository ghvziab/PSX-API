from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

def scraper():
    #excel file to store the data
    filepath = r'C:\Users\Wajiz.pk\Desktop\coding\psx extension\datapsx.xlsx'
    wb = load_workbook(filepath)
    ws = wb.active
    ws.title = "PSX Data"
    ws.delete_rows(1, ws.max_row)

    chrome_options = Options()
    chrome_options.add_argument('Mozilla/5.0 (Windows NT 10.0; Win64; x64)')

    driver = webdriver.Chrome(options=chrome_options)
    driver.get('https://dps.psx.com.pk')

    try:
        button1 = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//button[@class="tingle-modal__close"]'))
        )
        button1.click()
    except:
        pass

    #dropdowntable selector
    dropdown = driver.find_element(By.XPATH, '//select[@name="DataTables_Table_0_length"]')
    select = Select(dropdown)
    select.select_by_visible_text("All")
    driver.implicitly_wait(2)

    #finding tables->finding rows
    table = driver.find_element(By.XPATH, '//table[@class="tbl dataTable no-footer"]')
    rows = table.find_elements(By.TAG_NAME, "tr")

    #finding cells
    for row in rows[1:]:
        cols = row.find_elements(By.TAG_NAME, 'td')
        data = [col.text.strip() for col in cols]
        ws.append(data)
    wb.save(filepath)
